"""Excel output writer for group optimization results.

This module provides functionality to export group optimization results
to Excel files with detailed group information and statistical summaries.
"""

import os
from datetime import datetime
from typing import List, Dict, Any
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import GroupResult, Group, Participant


class ExcelWriter:
    """Handles Excel output generation for group optimization results."""
    
    def __init__(self):
        """Initialize the Excel writer."""
        pass
    
    def write_excel_output(self, result: GroupResult, output_path: str = None) -> str:
        """Write group optimization results to an Excel file.
        
        Args:
            result: The GroupResult containing optimized group assignments
            output_path: Optional custom output path. If None, generates timestamped filename
            
        Returns:
            The path to the created Excel file
            
        Raises:
            ValueError: If result is invalid or empty
            IOError: If file cannot be written
        """
        if not result or not result.groups:
            raise ValueError("Cannot write empty or invalid group result to Excel")
        
        # Generate output path if not provided
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"groups_{timestamp}.xlsx"
        
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create group sheets
        for group in result.groups:
            self._create_group_sheet(workbook, group)
        
        # Create summary sheet
        self._create_summary_sheet(workbook, result)
        
        # Save the workbook
        try:
            workbook.save(output_path)
            workbook.close()
        except Exception as e:
            raise IOError(f"Failed to save Excel file: {e}")
        
        return output_path
    
    def _create_group_sheet(self, workbook: Workbook, group: Group) -> None:
        """Create a worksheet for a single group.
        
        Args:
            workbook: The Excel workbook to add the sheet to
            group: The Group to create a sheet for
        """
        sheet_name = f"Group {group.id + 1}"
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Set up headers
        worksheet['A1'] = 'Participant Name'
        worksheet['B1'] = 'Score'
        worksheet['C1'] = 'Group Average'
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in ['A1', 'B1', 'C1']:
            worksheet[col].font = header_font
            worksheet[col].fill = header_fill
            worksheet[col].alignment = Alignment(horizontal='center')
        
        # Add member data
        group_average = group.average_score
        
        for i, member in enumerate(group.members, start=2):
            # Column A: Participant name (with advantage indicator)
            worksheet[f'A{i}'] = member.original_name
            
            # Column B: Individual score
            worksheet[f'B{i}'] = member.score
            worksheet[f'B{i}'].number_format = '0.0'
            
            # Column C: Group average (repeated for all rows)
            worksheet[f'C{i}'] = group_average
            worksheet[f'C{i}'].number_format = '0.00'
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add group summary at the bottom
        if group.members:
            summary_row = len(group.members) + 3
            
            worksheet[f'A{summary_row}'] = 'Group Summary:'
            worksheet[f'A{summary_row}'].font = Font(bold=True)
            
            worksheet[f'A{summary_row + 1}'] = f'Total Members: {len(group.members)}'
            worksheet[f'A{summary_row + 2}'] = f'Advantaged Members: {group.advantage_count}'
            worksheet[f'A{summary_row + 3}'] = f'Average Score: {group_average:.2f}'
    
    def _create_summary_sheet(self, workbook: Workbook, result: GroupResult) -> None:
        """Create a summary sheet with overall statistics.
        
        Args:
            workbook: The Excel workbook to add the sheet to
            result: The GroupResult to analyze
        """
        worksheet = workbook.create_sheet(title="Summary", index=0)
        
        # Collect all participants for overall statistics
        all_participants = []
        for group in result.groups:
            all_participants.extend(group.members)
        
        if not all_participants:
            worksheet['A1'] = 'No data available'
            return
        
        # Calculate summary statistics
        all_scores = [p.score for p in all_participants]
        summary_stats = self.create_summary_statistics(all_participants)
        
        # Set up the summary sheet layout
        row = 1
        
        # Title
        worksheet[f'A{row}'] = 'Group Balancer Summary Report'
        worksheet[f'A{row}'].font = Font(size=16, bold=True)
        row += 2
        
        # Generation timestamp
        worksheet[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        row += 2
        
        # Overall Statistics Section
        worksheet[f'A{row}'] = 'Overall Statistics'
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        stats_data = [
            ('Total Participants', len(all_participants)),
            ('Total Groups', len(result.groups)),
            ('Advantaged Participants', sum(1 for p in all_participants if p.has_advantage)),
            ('', ''),  # Empty row
            ('Highest Score', summary_stats['highest_score']),
            ('Lowest Score', summary_stats['lowest_score']),
            ('Average Score', summary_stats['average_score']),
            ('Median Score', summary_stats['median_score']),
            ('Standard Deviation', summary_stats['standard_deviation']),
            ('', ''),  # Empty row
            ('Score Variance Between Groups', result.score_variance)
        ]
        
        for label, value in stats_data:
            if label:  # Skip empty rows for labels
                worksheet[f'A{row}'] = label
                if isinstance(value, (int, float)) and value != '':
                    worksheet[f'B{row}'] = value
                    if isinstance(value, float):
                        worksheet[f'B{row}'].number_format = '0.00'
            row += 1
        
        # Group Distribution Section
        row += 1
        worksheet[f'A{row}'] = 'Group Distribution'
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Headers for group table
        worksheet[f'A{row}'] = 'Group'
        worksheet[f'B{row}'] = 'Members'
        worksheet[f'C{row}'] = 'Advantaged'
        worksheet[f'D{row}'] = 'Average Score'
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in ['A', 'B', 'C', 'D']:
            worksheet[f'{col}{row}'].font = header_font
            worksheet[f'{col}{row}'].fill = header_fill
            worksheet[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Group data
        for group in result.groups:
            worksheet[f'A{row}'] = f'Group {group.id + 1}'
            worksheet[f'B{row}'] = len(group.members)
            worksheet[f'C{row}'] = group.advantage_count
            worksheet[f'D{row}'] = group.average_score
            worksheet[f'D{row}'].number_format = '0.00'
            row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_statistics(self, participants: List[Participant]) -> Dict[str, float]:
        """Create summary statistics for a list of participants.
        
        Args:
            participants: List of participants to analyze
            
        Returns:
            Dictionary containing statistical measures
        """
        if not participants:
            return {
                'highest_score': 0.0,
                'lowest_score': 0.0,
                'average_score': 0.0,
                'median_score': 0.0,
                'standard_deviation': 0.0
            }
        
        scores = [p.score for p in participants]
        
        return {
            'highest_score': max(scores),
            'lowest_score': min(scores),
            'average_score': statistics.mean(scores),
            'median_score': statistics.median(scores),
            'standard_deviation': statistics.stdev(scores) if len(scores) > 1 else 0.0
        }
    
    def generate_timestamped_filename(self, base_name: str = "groups") -> str:
        """Generate a timestamped filename for Excel output.
        
        Args:
            base_name: Base name for the file (default: "groups")
            
        Returns:
            Timestamped filename with .xlsx extension
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{base_name}_{timestamp}.xlsx"