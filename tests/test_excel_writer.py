"""Tests for the Excel writer functionality."""

import pytest
import tempfile
import os
from hypothesis import given, strategies as st, settings, HealthCheck
from openpyxl import load_workbook
import statistics

from group_balancer.models import Participant, Group, GroupResult
from group_balancer.excel_writer import ExcelWriter


class TestExcelWriter:
    """Unit tests for the ExcelWriter class."""
    
    def test_write_excel_output_basic(self):
        """Test basic Excel output generation."""
        writer = ExcelWriter()
        
        # Create test data
        p1 = Participant("Alice", "Alice", 85.0, False)
        p2 = Participant("Bob", "Bob*", 90.0, True)
        group = Group(id=0, members=[p1, p2])
        
        result = GroupResult(
            groups=[group],
            score_variance=6.25,
            advantage_distribution=[1]
        )
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        
        try:
            # Write Excel output
            output_path = writer.write_excel_output(result, temp_path)
            assert output_path == temp_path
            assert os.path.exists(output_path)
            
            # Verify the file can be opened
            workbook = load_workbook(output_path)
            assert "Summary" in workbook.sheetnames
            assert "Group 1" in workbook.sheetnames
            workbook.close()
            
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
    
    def test_write_excel_output_empty_result(self):
        """Test error handling for empty results."""
        writer = ExcelWriter()
        
        # Test with None result
        with pytest.raises(ValueError, match="Cannot write empty or invalid group result"):
            writer.write_excel_output(None)
        
        # Test with empty groups
        empty_result = GroupResult(groups=[], score_variance=0.0, advantage_distribution=[])
        with pytest.raises(ValueError, match="Cannot write empty or invalid group result"):
            writer.write_excel_output(empty_result)
    
    def test_generate_timestamped_filename(self):
        """Test timestamped filename generation."""
        writer = ExcelWriter()
        
        filename = writer.generate_timestamped_filename()
        assert filename.startswith("groups_")
        assert filename.endswith(".xlsx")
        assert len(filename) == len("groups_YYYYMMDD_HHMMSS.xlsx")
        
        # Test custom base name
        custom_filename = writer.generate_timestamped_filename("custom")
        assert custom_filename.startswith("custom_")
        assert custom_filename.endswith(".xlsx")
    
    def test_create_summary_statistics(self):
        """Test summary statistics calculation."""
        writer = ExcelWriter()
        
        # Test with participants
        participants = [
            Participant("Alice", "Alice", 85.0, False),
            Participant("Bob", "Bob*", 90.0, True),
            Participant("Charlie", "Charlie", 75.0, False)
        ]
        
        stats = writer.create_summary_statistics(participants)
        
        assert stats['highest_score'] == 90.0
        assert stats['lowest_score'] == 75.0
        assert stats['average_score'] == 83.33333333333333
        assert stats['median_score'] == 85.0
        assert stats['standard_deviation'] > 0
        
        # Test with empty list
        empty_stats = writer.create_summary_statistics([])
        assert empty_stats['highest_score'] == 0.0
        assert empty_stats['lowest_score'] == 0.0
        assert empty_stats['average_score'] == 0.0
        assert empty_stats['median_score'] == 0.0
        assert empty_stats['standard_deviation'] == 0.0


class TestExcelWriterProperties:
    """Property-based tests for the Excel writer functionality."""
    
    @given(
        participant_data=st.lists(
            st.tuples(
                st.text(min_size=1, max_size=10, alphabet='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 '),
                st.floats(min_value=40.0, max_value=100.0, allow_nan=False, allow_infinity=False),
                st.booleans()
            ),
            min_size=4,
            max_size=8
        ),
        group_count=st.integers(min_value=2, max_value=3)
    )
    @settings(suppress_health_check=[HealthCheck.too_slow], deadline=None, max_examples=20)
    def test_excel_output_completeness_property(self, participant_data, group_count):
        """**Feature: group-balancer, Property 10: Excel output completeness**
        
        For any completed grouping, the Excel output should contain all groups with 
        member names, individual scores, and a summary sheet with highest score, 
        lowest score, average, median, and standard deviation.
        
        **Validates: Requirements 5.5, 5.6, 5.7**
        """
        # Skip if we don't have enough participants
        if len(participant_data) < group_count:
            return
        
        writer = ExcelWriter()
        
        # Create participants from the test data
        participants = []
        for i, (name, score, has_advantage) in enumerate(participant_data):
            original_name = f"{name}*" if has_advantage else name
            participants.append(Participant(name, original_name, score, has_advantage))
        
        # Create groups and distribute participants
        groups = []
        for i in range(group_count):
            groups.append(Group(id=i, members=[]))
        
        # Simple round-robin distribution for testing
        for i, participant in enumerate(participants):
            groups[i % group_count].add_member(participant)
        
        # Calculate variance for the result
        group_averages = [group.average_score for group in groups if group.members]
        score_variance = statistics.variance(group_averages) if len(group_averages) > 1 else 0.0
        advantage_distribution = [group.advantage_count for group in groups]
        
        # Create the result
        result = GroupResult(
            groups=groups,
            score_variance=score_variance,
            advantage_distribution=advantage_distribution
        )
        
        # Create temporary file for Excel output
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        
        try:
            # Write Excel output
            output_path = writer.write_excel_output(result, temp_path)
            assert os.path.exists(output_path), "Excel file should be created"
            
            # Load and verify the Excel file
            workbook = load_workbook(output_path)
            
            # Property 1: Should have a summary sheet
            assert "Summary" in workbook.sheetnames, "Summary sheet should exist"
            
            # Property 2: Should have one sheet per group
            for group in groups:
                expected_sheet_name = f"Group {group.id + 1}"
                assert expected_sheet_name in workbook.sheetnames, f"Sheet '{expected_sheet_name}' should exist"
            
            # Verify summary sheet content
            summary_sheet = workbook["Summary"]
            
            # Property 3: Summary sheet should contain required statistics
            summary_values = []
            for row in summary_sheet.iter_rows(values_only=True):
                if row[0] and row[1] is not None:
                    summary_values.append((str(row[0]), row[1]))
            
            summary_dict = dict(summary_values)
            
            # Check for required statistics
            all_scores = [p.score for p in participants]
            expected_stats = {
                'Highest Score': max(all_scores),
                'Lowest Score': min(all_scores),
                'Average Score': statistics.mean(all_scores),
                'Median Score': statistics.median(all_scores),
                'Standard Deviation': statistics.stdev(all_scores) if len(all_scores) > 1 else 0.0
            }
            
            for stat_name, expected_value in expected_stats.items():
                assert stat_name in summary_dict, f"Summary should contain '{stat_name}'"
                actual_value = summary_dict[stat_name]
                assert abs(actual_value - expected_value) < 0.01, f"{stat_name}: expected {expected_value}, got {actual_value}"
            
            # Property 4: Each group sheet should contain all member information
            for group in groups:
                if not group.members:  # Skip empty groups
                    continue
                    
                sheet_name = f"Group {group.id + 1}"
                group_sheet = workbook[sheet_name]
                
                # Get all data from the sheet
                sheet_data = list(group_sheet.iter_rows(values_only=True))
                
                # Property 4a: Should have headers
                headers = sheet_data[0] if sheet_data else []
                assert 'Participant Name' in headers, f"Group sheet should have 'Participant Name' header"
                assert 'Score' in headers, f"Group sheet should have 'Score' header"
                assert 'Group Average' in headers, f"Group sheet should have 'Group Average' header"
                
                # Property 4b: Should contain all group members
                member_names_in_sheet = []
                member_scores_in_sheet = []
                group_averages_in_sheet = []
                
                for row in sheet_data[1:]:  # Skip header row
                    if row[0]:  # If there's a name in the first column
                        member_names_in_sheet.append(str(row[0]))
                        if row[1] is not None:
                            member_scores_in_sheet.append(float(row[1]))
                        if row[2] is not None:
                            group_averages_in_sheet.append(float(row[2]))
                
                # Verify all members are present
                expected_names = [member.original_name for member in group.members]
                for expected_name in expected_names:
                    assert expected_name in member_names_in_sheet, f"Member '{expected_name}' should be in group sheet"
                
                # Property 4c: Should contain all individual scores
                expected_scores = [member.score for member in group.members]
                for expected_score in expected_scores:
                    assert any(abs(score - expected_score) < 0.01 for score in member_scores_in_sheet), f"Score {expected_score} should be in group sheet"
                
                # Property 4d: Group average should be consistent
                expected_average = group.average_score
                for avg in group_averages_in_sheet:
                    assert abs(avg - expected_average) < 0.01, f"Group average should be {expected_average}, got {avg}"
            
            # Property 5: File should be properly formatted Excel file
            assert len(workbook.sheetnames) == len(groups) + 1, f"Should have {len(groups)} group sheets + 1 summary sheet"
            
            workbook.close()
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.unlink(temp_path)