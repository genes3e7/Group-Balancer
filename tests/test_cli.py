"""Tests for the command-line interface functionality."""

import pytest
import tempfile
import os
import sys
from pathlib import Path
from unittest.mock import patch, MagicMock
from hypothesis import given, strategies as st
from openpyxl import Workbook

from group_balancer.cli import CLIHandler, CLIError, ArgumentParser
from group_balancer.models import Participant
from group_balancer.__main__ import main
from group_balancer.excel_reader import ExcelDataError


def _is_valid_positive_int(x):
    """Helper function to check if a string represents a valid positive integer."""
    try:
        return x.isdigit() and int(x) > 0
    except (ValueError, TypeError):
        return False


class TestArgumentParser:
    """Tests for command-line argument parsing."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.parser = ArgumentParser()
        
        # Create a temporary Excel file for testing
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Score'
        worksheet['A2'] = 'Alice'
        worksheet['B2'] = 85.0
        workbook.save(self.temp_file.name)
        workbook.close()
    
    def teardown_method(self):
        """Clean up test fixtures."""
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
    
    @given(group_count=st.integers(min_value=1, max_value=100))
    def test_group_count_consistency_property(self, group_count):
        """**Feature: group-balancer, Property 1: Group count consistency**
        
        For any valid group count argument, the system should create exactly 
        that number of groups.
        
        **Validates: Requirements 1.1**
        """
        # Test with valid group count and existing file
        args = [str(group_count), self.temp_file.name]
        
        parsed_args = self.parser.parse_args(args)
        
        # Verify the group count is preserved exactly
        assert parsed_args.group_count == group_count
        assert isinstance(parsed_args.group_count, int)
        assert parsed_args.group_count > 0
    
    @given(
        invalid_input=st.one_of(
            st.integers(max_value=0),  # Non-positive integers
            st.floats(),  # Float values
            st.text().filter(lambda x: not _is_valid_positive_int(x)),  # Non-numeric or non-positive strings
            st.just("-1"),  # Negative string
            st.just("0"),   # Zero string
            st.just("abc"), # Non-numeric string
            st.just(""),    # Empty string
        )
    )
    def test_invalid_input_rejection_property(self, invalid_input):
        """**Feature: group-balancer, Property 2: Invalid input rejection**
        
        For any invalid group count input (non-positive integers, non-numeric values), 
        the system should reject the input and display an appropriate error message.
        
        **Validates: Requirements 1.2**
        """
        # Convert input to string for command line argument
        if isinstance(invalid_input, (int, float)):
            if invalid_input <= 0:
                args = [str(invalid_input), self.temp_file.name]
            else:
                # Skip positive numbers that would be valid
                return
        else:
            args = [str(invalid_input), self.temp_file.name]
        
        # Should raise CLIError for invalid input
        with pytest.raises(CLIError):
            self.parser.parse_args(args)
    
    def test_missing_arguments(self):
        """Test error handling for missing arguments."""
        with pytest.raises(CLIError):
            self.parser.parse_args([])
        
        with pytest.raises(CLIError):
            self.parser.parse_args(["5"])  # Missing file argument
    
    def test_nonexistent_file(self):
        """Test error handling for nonexistent Excel files."""
        with pytest.raises(CLIError, match="Excel file not found"):
            self.parser.parse_args(["5", "nonexistent_file.xlsx"])
    
    def test_invalid_file_extension(self):
        """Test error handling for non-Excel files."""
        # Create a temporary text file
        temp_txt = tempfile.NamedTemporaryFile(suffix='.txt', delete=False)
        temp_txt.close()
        
        try:
            with pytest.raises(CLIError, match="File must be an Excel file"):
                self.parser.parse_args(["5", temp_txt.name])
        finally:
            os.unlink(temp_txt.name)
    
    def test_valid_arguments(self):
        """Test successful parsing of valid arguments."""
        args = ["5", self.temp_file.name]
        parsed_args = self.parser.parse_args(args)
        
        assert parsed_args.group_count == 5
        assert parsed_args.excel_file == self.temp_file.name
        assert parsed_args.output is None
    
    def test_valid_arguments_with_output(self):
        """Test successful parsing with output file specified."""
        args = ["3", self.temp_file.name, "--output", "results.xlsx"]
        parsed_args = self.parser.parse_args(args)
        
        assert parsed_args.group_count == 3
        assert parsed_args.excel_file == self.temp_file.name
        assert parsed_args.output == "results.xlsx"


class TestCLIHandler:
    """Tests for the main CLI handler."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.cli_handler = CLIHandler()
    
    def test_group_count_exceeds_participants(self):
        """Test validation when group count exceeds participant count."""
        participants = [
            Participant("Alice", "Alice", 85.0, False),
            Participant("Bob", "Bob", 90.0, False)
        ]
        
        # Should raise error when group count > participant count
        with pytest.raises(CLIError, match="Cannot create 5 groups with only 2 participants"):
            self.cli_handler.validate_group_count_against_participants(5, participants)
    
    def test_valid_group_count_for_participants(self):
        """Test validation passes for valid group count."""
        participants = [
            Participant("Alice", "Alice", 85.0, False),
            Participant("Bob", "Bob", 90.0, False),
            Participant("Charlie", "Charlie", 78.0, False),
            Participant("Diana", "Diana", 92.0, False)
        ]
        
        # Should not raise error for valid group count
        self.cli_handler.validate_group_count_against_participants(2, participants)
    
    def test_small_groups_warning(self, capsys):
        """Test warning for small groups."""
        participants = [
            Participant("Alice", "Alice", 85.0, False),
            Participant("Bob", "Bob", 90.0, False),
            Participant("Charlie", "Charlie", 78.0, False)
        ]
        
        # Should show warning but not raise error
        self.cli_handler.validate_group_count_against_participants(3, participants)
        
        captured = capsys.readouterr()
        assert "Warning" in captured.err
        assert "very small groups" in captured.err


class TestCLIIntegration:
    """Integration tests for the complete CLI workflow."""
    
    def setup_method(self):
        """Set up test fixtures for integration tests."""
        # Create a temporary Excel file with sample data
        self.temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.temp_file.close()
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Score'
        
        # Add sample participants with some having advantage status
        test_data = [
            ('Alice*', 85.0),
            ('Bob', 90.0),
            ('Charlie', 78.0),
            ('Diana*', 92.0),
            ('Eve', 88.0),
            ('Frank', 82.0)
        ]
        
        for i, (name, score) in enumerate(test_data, start=2):
            worksheet[f'A{i}'] = name
            worksheet[f'B{i}'] = score
        
        workbook.save(self.temp_file.name)
        workbook.close()
        
        # Create output directory for test files
        self.output_dir = tempfile.mkdtemp()
    
    def teardown_method(self):
        """Clean up test fixtures."""
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
        
        # Clean up output directory
        import shutil
        if os.path.exists(self.output_dir):
            shutil.rmtree(self.output_dir)
    
    def test_end_to_end_workflow_success(self, capsys):
        """Test complete end-to-end workflow with valid data."""
        output_file = os.path.join(self.output_dir, "test_output.xlsx")
        
        # Mock sys.argv to simulate command line arguments
        test_args = ['group-balancer', '2', self.temp_file.name, '--output', output_file]
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should return success (0)
        assert result == 0
        
        # Check console output
        captured = capsys.readouterr()
        assert "Loading participants" in captured.out
        assert "Loaded 6 participants" in captured.out
        assert "Creating 2 balanced groups" in captured.out
        assert "Group optimization completed successfully" in captured.out
        assert "GROUP BALANCER RESULTS" in captured.out
        assert "Excel output saved to" in captured.out
        
        # Check that output file was created
        assert os.path.exists(output_file)
    
    def test_end_to_end_workflow_with_timestamped_output(self, capsys):
        """Test workflow without specifying output file (uses timestamp)."""
        test_args = ['group-balancer', '3', self.temp_file.name]
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should return success (0)
        assert result == 0
        
        # Check console output mentions Excel output
        captured = capsys.readouterr()
        assert "Excel output saved to" in captured.out
        assert "groups_" in captured.out  # Timestamped filename
    
    def test_excel_file_not_found_error(self, capsys):
        """Test error handling when Excel file doesn't exist."""
        test_args = ['group-balancer', '2', 'nonexistent_file.xlsx']
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should return error (1)
        assert result == 1
        
        # Check error output
        captured = capsys.readouterr()
        assert "Error:" in captured.err
        assert "Excel file not found" in captured.err
    
    def test_invalid_group_count_error(self, capsys):
        """Test error handling for invalid group count."""
        test_args = ['group-balancer', '0', self.temp_file.name]
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should return error (1)
        assert result == 1
        
        # Check error output
        captured = capsys.readouterr()
        assert "Error:" in captured.err
        assert "Group count must be a positive integer" in captured.err
    
    def test_group_count_exceeds_participants_error(self, capsys):
        """Test error handling when group count exceeds participant count."""
        test_args = ['group-balancer', '10', self.temp_file.name]
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should return error (1)
        assert result == 1
        
        # Check error output
        captured = capsys.readouterr()
        assert "Error:" in captured.err
        assert "Cannot create 10 groups with only 6 participants" in captured.err
    
    def test_corrupted_excel_file_error(self, capsys):
        """Test error handling for corrupted Excel files."""
        # Create a corrupted file (not actually Excel format)
        corrupted_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        corrupted_file.write(b"This is not an Excel file")
        corrupted_file.close()
        
        try:
            test_args = ['group-balancer', '2', corrupted_file.name]
            
            with patch.object(sys, 'argv', test_args):
                result = main()
            
            # Should return error (1)
            assert result == 1
            
            # Check error output
            captured = capsys.readouterr()
            assert "Error:" in captured.err
            assert "Failed to load Excel data" in captured.err
            
        finally:
            os.unlink(corrupted_file.name)
    
    def test_excel_output_failure_graceful_degradation(self, capsys):
        """Test graceful degradation when Excel output fails."""
        # Use an invalid output path (directory that doesn't exist)
        invalid_output = "/nonexistent/directory/output.xlsx"
        test_args = ['group-balancer', '2', self.temp_file.name, '--output', invalid_output]
        
        with patch.object(sys, 'argv', test_args):
            result = main()
        
        # Should still return success (0) because console output worked
        assert result == 0
        
        # Check that warning was displayed but process continued
        captured = capsys.readouterr()
        assert "GROUP BALANCER RESULTS" in captured.out  # Console output still worked
        assert "Warning: Failed to create Excel output" in captured.err
        assert "Console output is still available above" in captured.err
    
    def test_keyboard_interrupt_handling(self, capsys):
        """Test handling of keyboard interrupt (Ctrl+C)."""
        test_args = ['group-balancer', '2', self.temp_file.name]
        
        # Mock the group optimizer to raise KeyboardInterrupt
        with patch.object(sys, 'argv', test_args):
            with patch('group_balancer.__main__.GroupOptimizer') as mock_optimizer:
                mock_optimizer.return_value.optimize_groups.side_effect = KeyboardInterrupt()
                
                result = main()
        
        # Should return error (1)
        assert result == 1
        
        # Check error output
        captured = capsys.readouterr()
        assert "Operation cancelled by user" in captured.err
    
    def test_unexpected_error_handling(self, capsys):
        """Test handling of unexpected errors."""
        test_args = ['group-balancer', '2', self.temp_file.name]
        
        # Mock the group optimizer to raise an unexpected error
        with patch.object(sys, 'argv', test_args):
            with patch('group_balancer.__main__.GroupOptimizer') as mock_optimizer:
                mock_optimizer.return_value.optimize_groups.side_effect = RuntimeError("Unexpected error")
                
                result = main()
        
        # Should return error (1)
        assert result == 1
        
        # Check error output - the error is wrapped in CLIError
        captured = capsys.readouterr()
        assert "Group optimization failed: Unexpected error" in captured.err
    
    def test_help_display(self, capsys):
        """Test that help is displayed correctly."""
        test_args = ['group-balancer', '--help']
        
        with patch.object(sys, 'argv', test_args):
            result = main()
            
            # Help should return with code 0
            assert result == 0
        
        # Check help output
        captured = capsys.readouterr()
        assert "Create balanced groups from participant data" in captured.out
        assert "group_count" in captured.out
        assert "excel_file" in captured.out
    
    def test_invalid_excel_data_error(self, capsys):
        """Test error handling for Excel files with invalid data."""
        # Create Excel file with invalid score data
        invalid_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        invalid_file.close()
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Score'
        worksheet['A2'] = 'Alice'
        worksheet['B2'] = 'invalid_score'  # Invalid score
        workbook.save(invalid_file.name)
        workbook.close()
        
        try:
            test_args = ['group-balancer', '2', invalid_file.name]
            
            with patch.object(sys, 'argv', test_args):
                result = main()
            
            # Should return error (1)
            assert result == 1
            
            # Check error output
            captured = capsys.readouterr()
            assert "Error:" in captured.err
            assert "Failed to load Excel data" in captured.err
            
        finally:
            os.unlink(invalid_file.name)
    
    def test_small_groups_warning_integration(self, capsys):
        """Test that small groups warning is displayed in integration."""
        # Create file with few participants
        small_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        small_file.close()
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Score'
        worksheet['A2'] = 'Alice'
        worksheet['B2'] = 85.0
        worksheet['A3'] = 'Bob'
        worksheet['B3'] = 90.0
        workbook.save(small_file.name)
        workbook.close()
        
        try:
            test_args = ['group-balancer', '2', small_file.name]
            
            with patch.object(sys, 'argv', test_args):
                result = main()
            
            # Should still succeed
            assert result == 0
            
            # Check warning output
            captured = capsys.readouterr()
            assert "Warning" in captured.err
            assert "very small groups" in captured.err
            
        finally:
            os.unlink(small_file.name)